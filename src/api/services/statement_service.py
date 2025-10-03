"""Service responsible for extracting statements from Excel workbooks."""

from __future__ import annotations

import logging
import shutil
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from types import SimpleNamespace
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from sec_downloader.utils import normalize_ticker

from ..data_models import FilingRecord, StatementRecord, StatementRow, StatementSheet
from .constants import STATEMENT_SHEET_NORMALIZATION, STATEMENT_TYPE_ALIASES
from .exceptions import DataRetrievalError
from .utils import normalize_date_bounds, normalize_statement_key

logger = logging.getLogger(__name__)


@dataclass(slots=True)
class _StatementCacheEntry:
    sheet: StatementSheet
    loaded_at: datetime
    ticker: str


class StatementRetrievalService:
    """Parses financial statements from generated Excel workbooks."""

    def __init__(
        self,
        filing_service,
        excel_dir: Path | str = Path("output"),
        cache_ttl: timedelta | int | float = timedelta(minutes=5),
    ) -> None:
        from .filing_service import FilingRetrievalService

        if not isinstance(filing_service, FilingRetrievalService):
            raise TypeError("filing_service must be a FilingRetrievalService instance")

        self.filings = filing_service
        self.excel_dir = Path(excel_dir)
        self._statement_cache: Dict[Tuple[Path, str], _StatementCacheEntry] = {}
        self._statement_index: Dict[str, Set[Tuple[Path, str]]] = {}
        if isinstance(cache_ttl, (int, float)):
            cache_ttl = timedelta(seconds=float(cache_ttl))
        if cache_ttl.total_seconds() < 0:
            raise ValueError("cache_ttl must be non-negative")
        self._cache_ttl = cache_ttl
        self._temp_root = filing_service.temp_root
        self.filings.register_download_listener(self._on_filing_downloaded)

    def get_statements_by_date_range(
        self,
        ticker: str,
        statement_type: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> List[StatementRecord]:
        normalized_ticker = normalize_ticker(ticker)
        canonical_statement = self._resolve_statement_type(statement_type)
        try:
            start_dt, end_dt = normalize_date_bounds(start_date, end_date)
        except ValueError as exc:
            raise DataRetrievalError(str(exc)) from exc

        logger.info(
            "Start statement date-range lookup for ticker=%s statement=%s start=%s end=%s",
            normalized_ticker,
            canonical_statement,
            start_dt,
            end_dt,
        )

        filings = self.filings.get_filings_by_date_range(
            normalized_ticker,
            start_date=start_date,
            end_date=end_date,
            document_type=None,
        )

        statements: List[StatementRecord] = []
        for filing in filings:
            ensured = self.filings.ensure_local_filing(filing) or filing
            record = self._build_statement_record(ensured, canonical_statement)
            if record:
                statements.append(record)

        logger.info(
            "Returning %d statements for ticker=%s",
            len(statements),
            normalized_ticker,
        )
        return statements

    def get_latest_statements(
        self,
        ticker: str,
        statement_type: str,
        limit: int = 4,
    ) -> List[StatementRecord]:
        normalized_ticker = normalize_ticker(ticker)
        canonical_statement = self._resolve_statement_type(statement_type)
        effective_limit = max(1, min(limit, 8))

        logger.info(
            "Start latest statement lookup for ticker=%s statement=%s limit=%s",
            normalized_ticker,
            canonical_statement,
            effective_limit,
        )

        filings = self.filings.get_latest_filings(
            normalized_ticker,
            limit=effective_limit,
            document_type=None,
        )

        statements: List[StatementRecord] = []
        for filing in filings:
            ensured = self.filings.ensure_local_filing(filing) or filing
            record = self._build_statement_record(ensured, canonical_statement)
            if record:
                statements.append(record)
            if len(statements) >= effective_limit:
                break

        logger.info(
            "Returning %d latest statements for ticker=%s",
            len(statements),
            normalized_ticker,
        )
        return statements

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------
    def _on_filing_downloaded(self, ticker: str) -> None:
        self._invalidate_statement_cache_for_ticker(ticker)

    def _build_statement_record(
        self, filing: FilingRecord, statement_type: str
    ) -> Optional[StatementRecord]:
        try:
            workbook_path = self._ensure_excel_file(filing)
        except DataRetrievalError as exc:
            logger.warning(
                "Unable to locate or generate Excel for %s: %s",
                filing.accession_number,
                exc,
            )
            return None

        sheet = self._load_statement_sheet(workbook_path, statement_type, filing.ticker)
        if not sheet:
            logger.info(
                "Statement %s not found in workbook %s",
                statement_type,
                workbook_path,
            )
            return None

        return StatementRecord(
            ticker=filing.ticker,
            statement_type=statement_type,
            filing_date=filing.filing_date,
            report_date=filing.report_date,
            form_type=filing.form_type,
            accession_number=filing.accession_number,
            workbook_path=workbook_path,
            sheet=sheet,
        )

    def _ensure_excel_file(self, filing: FilingRecord) -> Path:
        expected_path = self._build_expected_excel_path(filing)
        if expected_path.exists():
            return expected_path

        match = self._find_matching_excel(filing)
        if match:
            return match

        self._generate_excel_for_filing(filing, expected_path)
        if expected_path.exists():
            return expected_path

        raise DataRetrievalError(
            f"Excel workbook not available for accession {filing.accession_number}"
        )

    def _build_expected_excel_path(self, filing: FilingRecord) -> Path:
        ticker_dir = self.excel_dir / filing.ticker
        ticker_dir.mkdir(parents=True, exist_ok=True)
        date_source = filing.report_date or filing.filing_date
        date_part = date_source.strftime("%Y-%m-%d")
        accession_clean = filing.accession_number.replace("-", "")
        file_name = f"{filing.form_type}_{date_part}_{accession_clean}.xlsx"
        return ticker_dir / file_name

    def _find_matching_excel(self, filing: FilingRecord) -> Optional[Path]:
        ticker_dir = self.excel_dir / filing.ticker
        if not ticker_dir.exists():
            return None

        accession_clean = filing.accession_number.replace("-", "")
        for candidate in ticker_dir.glob(f"*{accession_clean}.xlsx"):
            return candidate
        return None

    def _generate_excel_for_filing(self, filing: FilingRecord, destination: Path) -> None:
        if not filing.local_path or not filing.primary_document:
            raise DataRetrievalError(
                f"Primary filing artifacts missing for accession {filing.accession_number}"
            )

        source_path = filing.local_path / filing.primary_document
        if not source_path.exists():
            raise DataRetrievalError(
                f"Primary document not found at {source_path}"
            )

        temp_dir = self._temp_root / f"{filing.ticker.lower()}_{filing.accession_number.replace('-', '')}"
        temp_dir.mkdir(parents=True, exist_ok=True)

        args = SimpleNamespace(
            filing=str(source_path),
            out=destination,
            one_period=False,
            periods=None,
            currency="USD",
            scale_none=False,
            include_disclosures=False,
            dump_role_map=None,
            label_style="terse",
            expand_dimensions=True,
            no_scale_hint=False,
            save_viewer_json=None,
            temp_dir=temp_dir,
            keep_temp=False,
            timeout=300,
            verbose=False,
            meta_links_candidates=[
                str(path)
                for path in [
                    filing.local_path / "MetaLinks.json",
                    filing.local_path / "metalink.json",
                ]
                if path.exists()
            ],
        )

        try:
            from render_viewer_to_xlsx import process_filing, validate_arguments

            validate_arguments(args)
            process_filing(args)
        except Exception as exc:  # pragma: no cover - pipeline invocation
            raise DataRetrievalError(
                f"Failed to generate Excel workbook: {exc}"
            ) from exc
        finally:
            if not args.keep_temp and temp_dir.exists():
                shutil.rmtree(temp_dir, ignore_errors=True)

    def _load_statement_sheet(
        self, workbook_path: Path, statement_type: str, ticker: str
    ) -> Optional[StatementSheet]:
        key = (workbook_path.resolve(), statement_type)
        entry = self._statement_cache.get(key)
        if entry and not self._is_cache_expired(entry.loaded_at):
            entry.loaded_at = datetime.utcnow()
            return entry.sheet

        sheet = self._parse_statement_sheet(workbook_path, statement_type)
        if sheet:
            self._cache_statement_sheet(key, sheet, ticker)
        return sheet

    def _parse_statement_sheet(
        self, workbook_path: Path, statement_type: str
    ) -> Optional[StatementSheet]:
        if not workbook_path.exists():
            return None

        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - corrupted workbook guard
            logger.warning("Failed to open workbook %s: %s", workbook_path, exc)
            return None

        try:
            canonical = statement_type
            target_sheet = None
            for name in workbook.sheetnames:
                normalized = normalize_statement_key(name)
                alias = STATEMENT_TYPE_ALIASES.get(normalized) or STATEMENT_SHEET_NORMALIZATION.get(
                    normalized
                )
                if alias == canonical:
                    target_sheet = workbook[name]
                    break

            if target_sheet is None:
                return None

            periods: List[str] = []
            column_indices: List[int] = []
            blank_streak = 0
            max_column = target_sheet.max_column or 1
            for col_idx in range(2, max_column + 1):
                header_value = target_sheet.cell(row=2, column=col_idx).value
                if header_value is None or str(header_value).strip() == "":
                    blank_streak += 1
                    if blank_streak >= 2 and periods:
                        break
                    continue
                blank_streak = 0
                periods.append(self._format_period_header(header_value))
                column_indices.append(col_idx)

            if not periods:
                return None

            rows: List[StatementRow] = []
            blank_row_streak = 0
            max_row = target_sheet.max_row or 3
            for row_idx in range(3, max_row + 1):
                label_value = target_sheet.cell(row=row_idx, column=1).value
                values: Dict[str, Optional[object]] = {}
                has_content = bool(label_value and str(label_value).strip())

                for period_label, col_idx in zip(periods, column_indices):
                    cell_value = target_sheet.cell(row=row_idx, column=col_idx).value
                    values[period_label] = cell_value
                    if cell_value not in (None, ""):
                        has_content = True

                if not has_content:
                    blank_row_streak += 1
                    if blank_row_streak >= 2 and rows:
                        break
                    continue

                blank_row_streak = 0
                label_text = str(label_value) if label_value is not None else ""
                rows.append(StatementRow(label=label_text, values=values))

            if not rows:
                return None

            return StatementSheet(sheet_name=target_sheet.title, periods=periods, rows=rows)
        finally:
            workbook.close()

    def _cache_statement_sheet(
        self, key: Tuple[Path, str], sheet: StatementSheet, ticker: str
    ) -> None:
        entry = _StatementCacheEntry(sheet=sheet, loaded_at=datetime.utcnow(), ticker=ticker)
        self._statement_cache[key] = entry
        self._statement_index.setdefault(ticker, set()).add(key)

    def _invalidate_statement_cache_for_ticker(self, ticker: str) -> None:
        keys = self._statement_index.pop(ticker, None)
        if not keys:
            return
        for key in keys:
            self._statement_cache.pop(key, None)

    def _is_cache_expired(self, loaded_at: datetime) -> bool:
        if self._cache_ttl.total_seconds() <= 0:
            return True
        return datetime.utcnow() - loaded_at > self._cache_ttl

    def _format_period_header(self, value: object) -> str:
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()  # type: ignore[return-value]
            except Exception:
                pass
        return str(value)

    def _resolve_statement_type(self, statement_type: str) -> str:
        key = normalize_statement_key(statement_type)
        canonical = STATEMENT_TYPE_ALIASES.get(key)
        if not canonical:
            raise DataRetrievalError(
                "Unsupported statementType '%s'. Allowed values: Income Statement, Balance Sheet, Cash Flows, Stockholders' Equity"
                % statement_type
            )
        return canonical
