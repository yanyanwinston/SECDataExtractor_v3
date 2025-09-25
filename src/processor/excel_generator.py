"""
Excel generator for SEC financial statements.
"""

import logging
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

from .data_models import Statement, ProcessingResult


logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generates Excel files from processed financial statements."""

    def __init__(self):
        """Initialize Excel generator."""
        pass

    def generate_excel(self, result: ProcessingResult, output_path: str,
                      single_period: bool = False) -> None:
        """
        Generate Excel file from processing result.

        Args:
            result: Processing result with statements
            output_path: Path for output Excel file
            single_period: Whether to show only the latest period

        Raises:
            Exception: If Excel generation fails
        """
        if not result.success or not result.statements:
            raise ValueError("Cannot generate Excel from failed processing result")

        try:
            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Add sheet for each statement
            for statement in result.statements:
                self._add_statement_sheet(wb, statement, single_period)

            # Add summary sheet if multiple statements
            if len(result.statements) > 1:
                self._add_summary_sheet(wb, result)

            # Save workbook
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_file)

            logger.info(f"Excel file generated successfully: {output_path}")

        except Exception as e:
            logger.error(f"Error generating Excel file: {e}")
            raise

    def _add_statement_sheet(self, wb: Workbook, statement: Statement,
                           single_period: bool) -> None:
        """
        Add a sheet for a single financial statement.

        Args:
            wb: Workbook object
            statement: Statement data
            single_period: Whether to show only latest period
        """
        # Create sheet with cleaned name
        sheet_name = self._clean_sheet_name(statement.short_name or statement.name)
        ws = wb.create_sheet(title=sheet_name)

        # Determine periods to show
        periods = statement.periods
        if single_period and periods:
            periods = [periods[0]]  # Show only the first (latest) period

        if not periods:
            logger.warning(f"No periods found for statement: {statement.name}")
            return

        # Set up headers
        self._write_headers(ws, periods)

        # Write data rows
        self._write_statement_rows(ws, statement, periods)

        # Apply formatting
        self._format_sheet(ws, len(periods))

    def _clean_sheet_name(self, name: str) -> str:
        """
        Clean sheet name for Excel compatibility.

        Args:
            name: Original sheet name

        Returns:
            Cleaned sheet name
        """
        # Remove/replace invalid characters
        invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
        cleaned = name

        for char in invalid_chars:
            cleaned = cleaned.replace(char, '_')

        # Limit length
        return cleaned[:31]

    def _write_headers(self, ws, periods: List) -> None:
        """
        Write column headers.

        Args:
            ws: Worksheet object
            periods: List of periods
        """
        # Row 1: Statement title (will be set later if needed)
        # Row 2: Period headers
        ws['A2'] = "Item"

        for i, period in enumerate(periods, start=2):
            col_letter = get_column_letter(i)
            # Use period label or end_date
            period_label = getattr(period, 'label', '') or getattr(period, 'end_date', '')
            ws[f'{col_letter}2'] = period_label

    def _write_statement_rows(self, ws, statement: Statement, periods: List) -> None:
        """
        Write statement data rows.

        Args:
            ws: Worksheet object
            statement: Statement data
            periods: List of periods to include
        """
        row_num = 3  # Start after headers

        for row in statement.rows:
            presentation_node = getattr(row, 'presentation_node', None)

            # Column A: Item label with indentation and styling
            label_cell = ws.cell(row=row_num, column=1)
            label_cell.value = row.label

            if presentation_node:
                indent_level = max(0, min(15, presentation_node.depth))
                label_cell.alignment = Alignment(indent=indent_level)

                if presentation_node.abstract:
                    label_cell.font = Font(bold=True)
                elif presentation_node.preferred_label_role:
                    role = presentation_node.preferred_label_role.lower()
                    if 'total' in role or 'subtotal' in role:
                        label_cell.font = Font(bold=True)
                        label_cell.border = Border(top=Side(style='thin'))
            else:
                # Legacy fallback using existing heuristics
                depth = getattr(row, 'depth', 0)
                if depth:
                    label_cell.alignment = Alignment(indent=max(0, min(15, depth)))
                if getattr(row, 'is_abstract', False):
                    label_cell.font = Font(bold=True)

            # Data columns
            for i, period in enumerate(periods, start=2):
                col_letter = get_column_letter(i)

                # Get cell value for this period
                period_key = getattr(period, 'label', '') or getattr(period, 'end_date', '')
                cell_data = row.cells.get(period_key)

                cell = ws.cell(row=row_num, column=i)

                if cell_data and cell_data.value is not None:
                    try:
                        if cell_data.raw_value is not None:
                            cell.value = float(cell_data.raw_value)
                        else:
                            cell.value = cell_data.value
                    except (ValueError, TypeError):
                        cell.value = cell_data.value

                    # Apply numeric formatting based on unit hints
                    if cell_data.raw_value is not None:
                        unit = (cell_data.unit or '').lower()
                        if 'usd' in unit:
                            cell.number_format = '#,##0.0_);(#,##0.0)'
                        elif 'shares' in unit:
                            cell.number_format = '#,##0'
                        elif unit in {'percent', '%'}:
                            cell.number_format = '0.00%'
                else:
                    cell.value = "—"

                # Highlight totals/subtotals using preferred label metadata
                if presentation_node and presentation_node.preferred_label_role:
                    role = presentation_node.preferred_label_role.lower()
                    if 'total' in role or 'subtotal' in role:
                        cell.border = Border(top=Side(style='thin'))

            row_num += 1

    def _format_sheet(self, ws, num_periods: int) -> None:
        """
        Apply formatting to the worksheet.

        Args:
            ws: Worksheet object
            num_periods: Number of period columns
        """
        # Define styles
        header_font = Font(bold=True, size=11)
        abstract_font = Font(bold=True, size=10)
        normal_font = Font(size=10)

        thin_border = Border(bottom=Side(style='thin'))

        # Header formatting (row 2)
        for col in range(1, num_periods + 2):  # +2 for label column
            cell = ws.cell(row=2, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Data rows formatting
        max_row = ws.max_row
        for row in range(3, max_row + 1):
            # Label column (A)
            label_cell = ws.cell(row=row, column=1)

            # Preserve explicit styling applied during row write; otherwise fall back
            if not label_cell.font or not label_cell.font.bold:
                label_cell.font = normal_font

            label_text = str(label_cell.value or "")
            if label_text and (label_cell.font and label_cell.font.bold):
                # Ensure bold rows maintain a separator if not already applied
                pass
            elif self._is_abstract_row(label_text):
                label_cell.font = abstract_font
                for col in range(1, num_periods + 2):
                    ws.cell(row=row, column=col).border = thin_border

            # Data columns
            for col in range(2, num_periods + 2):
                data_cell = ws.cell(row=row, column=col)
                data_cell.font = normal_font
                data_cell.alignment = Alignment(horizontal='right')

        # Set column widths
        ws.column_dimensions['A'].width = 50  # Item labels
        for col in range(2, num_periods + 2):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15

        # Freeze panes (freeze first column and header row)
        ws.freeze_panes = 'B3'

    def _is_abstract_row(self, label: str) -> bool:
        """
        Determine if a row is an abstract/section header.

        Args:
            label: Row label

        Returns:
            True if this appears to be an abstract row
        """
        if not label:
            return False

        # Simple heuristics for abstract rows
        abstract_indicators = [
            'total', 'assets', 'liabilities', 'equity', 'revenue', 'expenses',
            'income', 'cash flows', 'operating', 'investing', 'financing'
        ]

        label_lower = label.lower()

        # Check if it contains common abstract terms
        return any(indicator in label_lower for indicator in abstract_indicators)

    def _add_summary_sheet(self, wb: Workbook, result: ProcessingResult) -> None:
        """
        Add a summary sheet with metadata.

        Args:
            wb: Workbook object
            result: Processing result
        """
        ws = wb.create_sheet(title="Summary", index=0)

        # Company and filing info
        ws['A1'] = "Company Information"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = "Company:"
        ws['B3'] = result.company_name

        ws['A4'] = "Form Type:"
        ws['B4'] = result.form_type

        ws['A5'] = "Filing Date:"
        ws['B5'] = result.filing_date

        # Statements info
        ws['A7'] = "Financial Statements"
        ws['A7'].font = Font(bold=True, size=12)

        row = 9
        for i, statement in enumerate(result.statements, 1):
            ws[f'A{row}'] = f"{i}. {statement.name}"
            ws[f'B{row}'] = f"{len(statement.periods)} periods"
            ws[f'C{row}'] = f"{len(statement.rows)} line items"
            row += 1

        # Warnings if any
        if result.warnings:
            ws[f'A{row + 2}'] = "Warnings"
            ws[f'A{row + 2}'].font = Font(bold=True, size=12, color="FF6600")

            for i, warning in enumerate(result.warnings):
                ws[f'A{row + 4 + i}'] = f"• {warning}"

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
