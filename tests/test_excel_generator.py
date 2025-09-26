#!/usr/bin/env python3
"""Integration tests for ExcelGenerator output formatting."""

from pathlib import Path

import openpyxl

from src.processor.data_models import (
    Cell,
    Period,
    ProcessingResult,
    Row,
    Statement,
)
from src.processor.presentation_models import PresentationNode
from src.processor.excel_generator import ExcelGenerator


def _build_statement(name: str, period: Period) -> Statement:
    """Create a statement with presentation metadata for testing."""
    root_node = PresentationNode(
        concept="us-gaap:Assets",
        label="Assets",
        order=1,
        depth=0,
        abstract=True,
    )
    child_node = PresentationNode(
        concept="us-gaap:CashAndCashEquivalentsAtCarryingValue",
        label="Cash and Cash Equivalents",
        order=2,
        depth=1,
        abstract=False,
    )
    negative_node = PresentationNode(
        concept="us-gaap:OperatingLoss",
        label="Operating Loss",
        order=3.5,
        depth=1,
        abstract=False,
    )
    total_node = PresentationNode(
        concept="us-gaap:Assets",
        label="Total Assets",
        order=3,
        depth=1,
        abstract=False,
        preferred_label_role="totalLabel",
    )

    # Compose rows mirroring the presentation structure
    rows = []
    for node, raw_value in (
        (root_node, None),
        (child_node, 1234000.0),
        (negative_node, -500000.0),
        (total_node, 1234000.0),
    ):
        cells = {
            period.label: Cell(
                value="1234" if raw_value is None else str(int(raw_value)),
                raw_value=raw_value,
                unit="usd",
                decimals=-3,
                period=period.label,
            )
        }

        row = Row(
            label=node.label,
            concept=node.concept,
            is_abstract=node.abstract,
            depth=node.depth,
            cells=cells,
        )
        row.presentation_node = node
        rows.append(row)

    return Statement(
        name=name,
        short_name=name,
        periods=[period],
        rows=rows,
    )


def test_excel_generator_applies_presentation_formatting(tmp_path):
    period = Period(label="2023-09-30", end_date="2023-09-30", instant=True)
    statement = _build_statement("Balance Sheet", period)

    result = ProcessingResult(
        statements=[statement],
        company_name="Example Co",
        filing_date="2023-09-30",
        form_type="10-K",
        success=True,
    )

    output_path = Path(tmp_path) / "presentation.xlsx"
    ExcelGenerator().generate_excel(result, str(output_path))

    wb = openpyxl.load_workbook(output_path)
    ws = wb[statement.short_name]

    # Row ordering: headers on row 2, data starts at row 3
    header_row = 2
    first_data_row = header_row + 1

    # Header row should include the "Item" label
    assert ws.cell(row=header_row, column=1).value == "Item"
    assert ws.cell(row=header_row, column=2).value == period.label

    # Index rows by label for easier assertions
    label_rows = {
        ws.cell(row=row_idx, column=1).value: row_idx
        for row_idx in range(first_data_row, ws.max_row + 1)
    }

    # Root abstract row should be bold with no indentation
    root_cell = ws.cell(row=label_rows["Assets"], column=1)
    assert root_cell.font.bold is True
    assert root_cell.alignment.indent == 0

    # Child row should inherit depth-based indentation
    child_cell = ws.cell(row=label_rows["Cash and Cash Equivalents"], column=1)
    assert child_cell.alignment.indent == 1
    assert child_cell.font.bold is False

    # Negative row uses currency formatting and remains right aligned
    loss_value_cell = ws.cell(row=label_rows["Operating Loss"], column=2)
    assert loss_value_cell.value == -500000.0
    assert loss_value_cell.number_format == "#,##0.0_);(#,##0.0)"

    # Total row should have a top border and bold label
    total_label_cell = ws.cell(row=label_rows["Total Assets"], column=1)
    total_value_cell = ws.cell(row=label_rows["Total Assets"], column=2)
    assert total_label_cell.font.bold is True
    assert total_value_cell.border.top.style == "thin"
    assert total_value_cell.number_format == "#,##0.0_);(#,##0.0)"

    # Data cell should be numeric after raw_value handling
    assert (
        ws.cell(row=label_rows["Cash and Cash Equivalents"], column=2).value
        == 1234000.0
    )


def test_excel_generator_creates_summary_sheet(tmp_path):
    period = Period(label="2023-12-31", end_date="2023-12-31", instant=True)
    statements = [
        _build_statement("Balance Sheet", period),
        _build_statement("Income Statement", period),
    ]

    result = ProcessingResult(
        statements=statements,
        company_name="Example Co",
        filing_date="2023-12-31",
        form_type="10-K",
        success=True,
    )

    output_path = Path(tmp_path) / "summary.xlsx"
    ExcelGenerator().generate_excel(result, str(output_path))

    wb = openpyxl.load_workbook(output_path)

    # Summary sheet inserted at index 0
    assert wb.sheetnames[0] == "Summary"
    summary_ws = wb["Summary"]

    assert summary_ws["A3"].value == "Company:"
    assert summary_ws["B3"].value == result.company_name
    assert summary_ws["A7"].value == "Financial Statements"

    # Ensure each statement is listed with counts
    listed_statements = {summary_ws[f"A{row}"].value for row in range(9, 11)}
    assert any("Balance Sheet" in entry for entry in listed_statements if entry)
    assert any("Income Statement" in entry for entry in listed_statements if entry)

    # Primary statement sheets should exist alongside the summary
    sheet_names = set(wb.sheetnames)
    assert {"Balance Sheet", "Income Statement"}.issubset(sheet_names)
